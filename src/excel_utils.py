import os
import re
import shutil
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

def apply_formatting(source_path: str, source_sheet_name: str, target_path: str):
    """
    Copies column widths and wrap_text ALIGNMENT from source to target.
    Does NOT copy fonts or colors (to preserve our Analysis highlighting).
    """
    try:
        # Load Source
        wb_source = openpyxl.load_workbook(source_path, read_only=False, data_only=True)
        if source_sheet_name and source_sheet_name in wb_source.sheetnames:
            ws_source = wb_source[source_sheet_name]
        else:
            ws_source = wb_source.active

        # Load Target
        wb_target = openpyxl.load_workbook(target_path)
        ws_target = wb_target.active # output usually has only one sheet

        # 1. Copy Column Widths
        # We assume the columns map 1:1 for at least the master columns (A, B, C...)
        # Iterate over source columns
        for col_char in ['A', 'B', 'C', 'D', 'E', 'F']:
            # Safe limit, we mostly care about Description (B) and S.No (A)
            if col_char in ws_source.column_dimensions:
                source_dim = ws_source.column_dimensions[col_char]
                if source_dim.customWidth: # Only copy custom widths
                    ws_target.column_dimensions[col_char].width = source_dim.width
        
        # 2. Inspect Description Column (B) for Wrap Text
        # We check the first few rows to see if wrap is enabled
        wrap_enabled = False
        for i in range(1, 20):
            cell = ws_source[f'B{i}']
            if cell.alignment and cell.alignment.wrap_text:
                wrap_enabled = True
                break
        
        if wrap_enabled:
            # Apply wrap text to Column B in target
            # Note: We must apply to every cell in the column or setup a style
            for row in ws_target.iter_rows(min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

        # 3. Auto-adjust removed as per user request
        # We leave Vendor columns (Rate/Amount) to default Excel width (or Pandas default)


        wb_target.save(target_path)
        print(f"Formatting applied to {target_path}")

    except Exception as e:
        print(f"Formatting error: {e}")

from openpyxl.styles import Border, Side, Alignment

def overlay_analysis_on_template(ref_path: str, output_path: str, final_df: pd.DataFrame, start_row: int, sheet_name=None):
    """
    Overlays the finalized dataframe onto a copy of the reference file.
    Preserves header metadata (rows < start_row) and footer (implicitly, if we don't overwrite).
    Actually, footers might be overwritten if the new table is longer.
    But usually we want to 'insert' or just overwrite the table area.
    
    If the new table is LONGER than the original, we just write over whatever was below.
    If the footer was immediately below the original table, it might be overwritten.
    Ideal solution: Insert rows? But that breaks merged cells often.
    
    For now, we overwrite starting at start_row (header).
    The user accepted "Template Overlay".
    """
    import shutil
    import pandas as pd
    
    # 1. Clone Reference
    shutil.copy(ref_path, output_path)
    
    try:
        wb = openpyxl.load_workbook(output_path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            
        # 2. Write Headers
        # start_row is 0-indexed from pandas read, but Excel is 1-indexed.
        # If openpyxl load is 1-based.
        # Header row in Excel = start_row + 1
        excel_header_row = start_row + 1
        
        # Write Column Headers
        for c_idx, col_name in enumerate(final_df.columns, 1):
            cell = ws.cell(row=excel_header_row, column=c_idx)
            cell.value = col_name
            # Optional: Apply header style to new columns if needed
            
        # 3. Write Data
        # Iterating dataframe rows
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Pre-calculate ranges to unmerge safely? 
        # Writing cell-by-cell and unmerging on the fly is tricky because unmerge invalidates the cell object?
        # No, unmerge just changes metadata.
        
        
        for r_idx, row in enumerate(final_df.itertuples(index=False), 1):
            excel_row = excel_header_row + r_idx
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=excel_row, column=c_idx)
                
                # Check for MergedCell
                if type(cell).__name__ == 'MergedCell':
                    for rng in ws.merged_cells.ranges:
                        if cell.coordinate in rng:
                            ws.unmerge_cells(str(rng))
                            cell = ws.cell(row=excel_row, column=c_idx)
                            break
                
                try:
                    # Sanitize NaN values to avoid "nan" string in Excel
                    if pd.isna(value) or str(value).lower() == 'nan':
                        value = ""
                        
                    cell.value = value
                except AttributeError:
                    print(f"Skipping write to {cell.coordinate} - ReadOnly/Merged")
                
                # Apply simple border
                cell.border = border

                # --- Enforce Formatting ---
                # We identify columns by header logic or dynamic lookup if possible
                # But here we are iterating rows.
                # Let's rely on the column structure from main.py: s.no, description, ...
                # Or check the column header if we tracked it.
                
                # Heuristic: Description is usually the longest string. 
                # Or use proper index checking:
                # We know final_report columns. We can find the index of 'description'
                # But inside this loop we don't have easy modification of the outer scope variables without re-reading
                # Let's assume description is column 2 (standard) OR check matching known headers.
                
                is_desc = (c_idx == 2) # Fallback default
                # Better: Check column name from dataframe if possible. 
                # We have 'final_df.columns'. c_idx is 1-based.
                # col_name = final_df.columns[c_idx - 1]
                # if 'description' in str(col_name).lower(): ...
                
                current_col_name = str(final_df.columns[c_idx - 1]).lower()
                is_desc = 'description' in current_col_name
                
                if is_desc:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Auto-Calculate Row Height using Actual Column Width
                    text_val = str(value) if value else ""
                    
                    # Get actual column width
                    # col_letter from c_idx
                    col_char = get_column_letter(c_idx)
                    col_width = ws.column_dimensions[col_char].width
                    
                    # If default/None, assume standard ~10-12 chars? Or usually ~8.43
                    if not col_width:
                        col_width = 10.0
                    
                    # Effective width for chars is slightly less than column width unit usually?
                    # Actually, openpyxl 'width' is roughly number of '0' characters in default font.
                    # We accept it as chars_per_line approximation.
                    chars_per_line = max(int(col_width), 10) # safety floor
                    
                    num_lines = (len(text_val) // chars_per_line) + 1
                    
                    # Only increase height if we have multiple lines
                    if num_lines > 1:
                        # Standard height ~15 per line.
                        # Maybe 14 to be tighter as requested ("more than required")
                        new_h = num_lines * 14.0
                        
                        current_h = ws.row_dimensions[excel_row].height
                        if current_h is None or new_h > current_h:
                             ws.row_dimensions[excel_row].height = new_h
                             
                else:
                    # Others: Top Align for tidiness
                    cell.alignment = Alignment(vertical='top')
                
        # 5. Auto-width for Vendor & Analysis Columns
        # We want to fit 'Rate_VendorName', 'Amount_VendorName', and 'Lowest_Vendor'/'Highest_Vendor'
        for c_idx in range(1, ws.max_column + 1):
             cell = ws.cell(row=excel_header_row, column=c_idx)
             val = str(cell.value)
             if val.startswith('Rate_') or val.startswith('Amount_') or val in ['Lowest_Vendor', 'Highest_Vendor']:
                 col_letter = get_column_letter(c_idx)
                 # Length + padding
                 ws.column_dimensions[col_letter].width = len(val) + 4

        # 4. Save
        wb.save(output_path)
        print(f"Overlay complete: {output_path}")
        
    except Exception as e:
        print(f"Overlay Error: {e}")


from openpyxl.styles import PatternFill

def apply_openpyxl_highlighting(file_path: str, metadata: dict, start_row: int, columns: list, sheet_name=None):
    """
    Applies Green/Red highlighting to the generated Excel file using OpenPyXL.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") # LightGreen
        red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")   # LightCoral
        
        # Header in Excel is start_row + 1. Data starts at start_row + 2
        data_start_row = start_row + 2
        
        min_max_data = metadata.get('min_max', {})
        
        # Map column names to indices for fast lookup
        # columns is the list of df columns
        rate_col_indices = []
        for idx, col in enumerate(columns):
            if str(col).startswith("Rate_"):
                # Excel index is idx + 1
                rate_col_indices.append(idx + 1)
                
        if not rate_col_indices:
            return

        # Iterate rows in metadata
        for row_idx, stats in min_max_data.items():
            # final_report row_index corresponds to excel row: data_start_row + row_idx (if index is 0-based sequential)
            # BUT 'row_idx' from metadata comes from final_report.index. 
            # If final_report was just a copy of ref_df reset_index, it's 0, 1, 2...
            # Yes, main.py does report_df = ref_df.copy() and no filtering that changes index numbers relative to position 
            # (unless we filtered rows, but we didn't).
            
            excel_r = data_start_row + row_idx
            
            r_min = stats['min']
            r_max = stats['max']
            
            for c_idx in rate_col_indices:
                cell = ws.cell(row=excel_r, column=c_idx)
                val = cell.value
                
                try:
                    if val is not None and float(val) == r_min:
                        cell.fill = green_fill
                    elif val is not None and float(val) == r_max:
                        cell.fill = red_fill
                except:
                    pass
                    
        wb.save(file_path)
        print(f"Highlighting applied to {file_path}")
        
    except Exception as e:
        print(f"Highlighting Error: {e}")

def clean_excel_file(file_path: str):
    """
    Strips definedNames from the workbook.xml of an .xlsx file.
    This fixes openpyxl crashes on files saved from older formats containing broken named ranges.
    """
    if not str(file_path).lower().endswith('.xlsx'):
        return
        
    try:
        temp_dir = str(file_path) + "_temp_unzip"
        os.makedirs(temp_dir, exist_ok=True)
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        workbook_xml_path = os.path.join(temp_dir, 'xl', 'workbook.xml')
        if os.path.exists(workbook_xml_path):
            with open(workbook_xml_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Remove definedNames block
            if '<definedNames>' in content:
                cleaned_content = re.sub(r'<definedNames>.*?</definedNames>', '', content, flags=re.DOTALL)
                
                with open(workbook_xml_path, 'w', encoding='utf-8') as f:
                    f.write(cleaned_content)
                
                # Repack only if modified
                with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            abs_path = os.path.join(root, file)
                            rel_path = os.path.relpath(abs_path, temp_dir)
                            zipf.write(abs_path, rel_path)
                            
        shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception as e:
        print(f"Failed to clean excel file {file_path}: {e}")

